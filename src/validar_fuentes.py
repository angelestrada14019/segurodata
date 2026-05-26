"""
Valida las 20 fuentes del proyecto y genera fuentes_validadas.xlsx con 4 hojas:
  Hoja 1 — Catálogo      : metadata completa de las 20 fuentes
  Hoja 2 — Geoespaciales : URLs, CRS y cómo hacer spatial join con UPZ
  Hoja 3 — Descartadas   : 13 fuentes descartadas con razón y alternativa
  Hoja 4 — Frecuencias   : semáforo de calidad (verde/amarillo/rojo) + impacto si stale

Uso:
  python src/validar_fuentes.py
  → Excel guardado en fuentes_validadas.xlsx
"""
from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).parent))
from etl import check_url

# ─────────────────────────────────────────────────────────────────────────────
# Catálogo de 20 fuentes validadas
# ─────────────────────────────────────────────────────────────────────────────

FUENTES = [
    {
        "id": 1,
        "nombre": "Incidentes Tramitados C4 — NUSE 123",
        "organizacion": "Secretaría Distrital de Seguridad, Convivencia y Justicia",
        "portal": "datosabiertos.bogota.gov.co",
        "plataforma": "CKAN Datastore",
        "tipo_acceso": "API filtrada (sin descarga completa)",
        "endpoint_url": "https://datosabiertos.bogota.gov.co/api/3/action/datastore_search?resource_id=30d65a8b-d0ed-4e95-977e-0d7cc2ea89ef",
        "resource_id": "30d65a8b-d0ed-4e95-977e-0d7cc2ea89ef",
        "frecuencia_declarada": "Mensual",
        "ultima_actualizacion_real": "20/04/2026",
        "periodo_cobertura": "Enero 2015 – Marzo 2026",
        "granularidad_temporal": "Mensual agregado",
        "granularidad_geografica": "UPZ (COD_UPZ formato 'UPZ99')",
        "variables_clave": "ANIO, MES, TIPO_INCIDENTE, COD_UPZ, UPZ, CANT_INCIDENTES",
        "por_que_importante": "Proxy de inseguridad percibida. Feature clave: ratio NUSE/delitos_formales_UPZ para cuantificar subregistro.",
        "valor_modelo": "Feature de subregistro — responde la pregunta difícil del jurado",
        "licencia": "CC BY-SA 4.0",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "Datos agregados (no registro individual). Sin coordenadas, solo código UPZ.",
        "volumen_estimado": "1,068,340 registros",
    },
    {
        "id": 2,
        "nombre": "Delito de Alto Impacto. Bogotá D.C.",
        "organizacion": "Secretaría Distrital de Seguridad, Convivencia y Justicia",
        "portal": "datosabiertos.bogota.gov.co",
        "plataforma": "CKAN — descarga de archivo",
        "tipo_acceso": "Descarga GeoJSON ZIP",
        "endpoint_url": "https://datosabiertos.bogota.gov.co/dataset/7b270013-42ca-436b-9c1e-3bcb7d280c6b/resource/aba0e25d-d407-45f4-9a98-327493b538bd/download/dai_geojson.zip",
        "resource_id": "aba0e25d-d407-45f4-9a98-327493b538bd",
        "frecuencia_declarada": "Semestral",
        "ultima_actualizacion_real": "20/04/2026",
        "periodo_cobertura": "Enero 2018 – Marzo 2026",
        "granularidad_temporal": "Registro individual por delito",
        "granularidad_geografica": "Punto georreferenciado + UPZ",
        "variables_clave": "tipologia_delito, lat, lon, fecha, UPZ, localidad",
        "por_que_importante": "FUENTE PRINCIPAL de crimen georreferenciado. Permite spatial join con UPZ y análisis de hotspots.",
        "valor_modelo": "Variable objetivo del modelo (Y). Registros geolocalizados.",
        "licencia": "CC BY-SA 4.0",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "Solo descarga de archivo, no API tabular. Semestral = posible desfase 2do semestre.",
        "volumen_estimado": ">500K registros (estimado)",
    },
    {
        "id": 3,
        "nombre": "HOMICIDIO — Policía Nacional de Colombia",
        "organizacion": "Policía Nacional de Colombia",
        "portal": "www.datos.gov.co",
        "plataforma": "Socrata",
        "tipo_acceso": "API Socrata (sodapy)",
        "endpoint_url": "https://www.datos.gov.co/resource/m8fd-ahd9.json",
        "resource_id": "m8fd-ahd9",
        "frecuencia_declarada": "Mensual",
        "ultima_actualizacion_real": "Datos Abril 2026",
        "periodo_cobertura": "Desde 2003",
        "granularidad_temporal": "Mensual",
        "granularidad_geografica": "Municipal",
        "variables_clave": "fecha_hecho, departamento, municipio, zona, sexo, arma_medio, cantidad",
        "por_que_importante": "Benchmarking Bogotá vs promedio nacional. Tendencia histórica 20+ años.",
        "valor_modelo": "Contexto comparativo — no feature directa (nivel municipal)",
        "licencia": "Datos abiertos públicos",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "Solo nivel municipal (no UPZ). Sin App Token: límite 1000 filas/request.",
        "volumen_estimado": ">1M registros",
    },
    {
        "id": 4,
        "nombre": "HURTO PERSONAS — Policía Nacional de Colombia",
        "organizacion": "Policía Nacional de Colombia",
        "portal": "www.datos.gov.co",
        "plataforma": "Socrata",
        "tipo_acceso": "API Socrata (sodapy)",
        "endpoint_url": "https://www.datos.gov.co/resource/4rxi-8m8d.json",
        "resource_id": "4rxi-8m8d",
        "frecuencia_declarada": "Mensual",
        "ultima_actualizacion_real": "Datos Abril 2026",
        "periodo_cobertura": "Multianual",
        "granularidad_temporal": "Mensual",
        "granularidad_geografica": "Municipal",
        "variables_clave": "fecha_hecho, departamento, municipio, cantidad",
        "por_que_importante": "Hurto es el delito más frecuente en Bogotá. Complementa homicidios para análisis multitipo.",
        "valor_modelo": "Contexto comparativo — no feature directa (nivel municipal)",
        "licencia": "Datos abiertos públicos",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "Solo nivel municipal. Columnas más limitadas que el dataset de homicidios.",
        "volumen_estimado": ">500K registros",
    },
    {
        "id": 5,
        "nombre": "Open-Meteo — Climate API",
        "organizacion": "Open-Meteo (servicio internacional libre)",
        "portal": "open-meteo.com",
        "plataforma": "REST API propia",
        "tipo_acceso": "REST API gratuita sin clave",
        "endpoint_url": "https://archive-api.open-meteo.com/v1/archive?latitude=4.6097&longitude=-74.0817&start_date=2020-01-01&end_date=2020-01-02&hourly=temperature_2m",
        "resource_id": "N/A",
        "frecuencia_declarada": "Horaria / tiempo real",
        "ultima_actualizacion_real": "Siempre actual",
        "periodo_cobertura": "Desde 1940 (API de archivo)",
        "granularidad_temporal": "Horaria o diaria (configurable)",
        "granularidad_geografica": "Bogotá (lat 4.6097, lon -74.0817)",
        "variables_clave": "temperature_2m (°C), precipitation (mm), windspeed_10m, relativehumidity_2m",
        "por_que_importante": "Feature temporal crítica. Lluvia→reduce hurto. Temperatura alta→aumenta riñas. Única fuente en tiempo real.",
        "valor_modelo": "Feature climática directa. Necesaria para predicciones hora por hora.",
        "licencia": "CC BY 4.0",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "Rate limit 10K req/día (no aplica uso normal). Histórico en endpoint /archive, pronóstico en /forecast.",
        "volumen_estimado": "Ilimitado (API)",
    },
    {
        "id": 6,
        "nombre": "UPZ Shapefile — Unidad de Planeamiento Local (IDECA)",
        "organizacion": "IDECA",
        "portal": "datosabiertos.bogota.gov.co",
        "plataforma": "CKAN — descarga directa",
        "tipo_acceso": "Descarga GeoJSON / SHP ZIP",
        "endpoint_url": "https://datosabiertos.bogota.gov.co/dataset/808582fc-ffc8-4649-8428-7e1fd8d3820c/resource/a5c8c591-0708-420f-8eb7-9f3147e21c40/download/unidadplaneamientolocal.json",
        "resource_id": "a5c8c591-0708-420f-8eb7-9f3147e21c40",
        "frecuencia_declarada": "Según necesidad",
        "ultima_actualizacion_real": "20/02/2025",
        "periodo_cobertura": "Actual",
        "granularidad_temporal": "Estática (infraestructura)",
        "granularidad_geografica": "Polígono por UPZ",
        "variables_clave": "codigo_upz, nombre, geometry (POLYGON), area, localidad",
        "por_que_importante": "BASE ESPACIAL del modelo. Sin esta capa no hay spatial joins ni mapa del dashboard.",
        "valor_modelo": "Capa geoespacial base — sin esta no hay modelo UPZ.",
        "licencia": "CC BY 4.0",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "Verificar que el GeoJSON tiene exactamente 112 polígonos (no 117 UPL).",
        "volumen_estimado": "112 polígonos",
    },
    {
        "id": 7,
        "nombre": "Cuadrante de Policía. Bogotá D.C.",
        "organizacion": "MEBOG / Policía Metropolitana de Bogotá",
        "portal": "datosabiertos.bogota.gov.co",
        "plataforma": "CKAN — descarga directa",
        "tipo_acceso": "Descarga GeoJSON / SHP ZIP",
        "endpoint_url": "https://datosabiertos.bogota.gov.co/dataset/cuadrantes-de-policia-bogota-d-c",
        "resource_id": "f0ad2ee3-bfd0-4825-9b31-bff9041649fa",
        "frecuencia_declarada": "Anual",
        "ultima_actualizacion_real": "02/03/2026",
        "periodo_cobertura": "Actual",
        "granularidad_temporal": "Estática (infraestructura)",
        "granularidad_geografica": "Polígono por cuadrante",
        "variables_clave": "codigo_cuadrante, geometry, area, localidad",
        "por_que_importante": "Feature de cobertura policial: cuadrantes/km² en cada UPZ. Input de capa prescriptiva.",
        "valor_modelo": "Feature urbanística: densidad de cuadrantes por UPZ.",
        "licencia": "Pública",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "Solo geometría, sin datos de actividad policial (esos son NUSE 123 y Delito Alto Impacto).",
        "volumen_estimado": "~1200 cuadrantes",
    },
    {
        "id": 8,
        "nombre": "Alumbrado Público, Bogotá D.C. (UAESP)",
        "organizacion": "UAESP",
        "portal": "datosabiertos.bogota.gov.co",
        "plataforma": "CKAN — descarga directa",
        "tipo_acceso": "Descarga GPKG / GeoJSON",
        "endpoint_url": "https://datosabiertos.bogota.gov.co/dataset/luminarias_upz-bogota-d-c",
        "resource_id": "luminarias_upz-bogota-d-c",
        "frecuencia_declarada": "Trimestral",
        "ultima_actualizacion_real": "19/05/2022  STALE 4 años",
        "periodo_cobertura": "2022",
        "granularidad_temporal": "Punto en el tiempo (infraestructura)",
        "granularidad_geografica": "Por localidad y UPZ",
        "variables_clave": "luminarias_led, luminarias_sodio, luminarias_total, localidad, upz",
        "por_que_importante": "Feature urbanística de iluminación declarada. Complementada por FUENTE 20 (VIIRS iluminación real).",
        "valor_modelo": "Feature de infraestructura de iluminación declarada.",
        "licencia": "Uso abierto",
        "estado_url": None,
        "stale": "SÍ ⚠️",
        "advertencias_calidad": "4 años stale — documentar en Notebook 05. Infraestructura cambia lentamente. Usar VIIRS (FUENTE 20) como complemento de iluminación real.",
        "volumen_estimado": "112 filas (una por UPZ)",
    },
    {
        "id": 9,
        "nombre": "Estaciones Troncales de TransMilenio",
        "organizacion": "TransMilenio S.A.",
        "portal": "datosabiertos.bogota.gov.co",
        "plataforma": "CKAN",
        "tipo_acceso": "Descarga directa / ArcGIS REST",
        "endpoint_url": "https://datosabiertos.bogota.gov.co/dataset/9be8b6fb-8059-492f-a866-4a1ac031c502",
        "resource_id": "9be8b6fb-8059-492f-a866-4a1ac031c502",
        "frecuencia_declarada": "Según necesidad",
        "ultima_actualizacion_real": "10/07/2025",
        "periodo_cobertura": "Actual",
        "granularidad_temporal": "Estática (infraestructura)",
        "granularidad_geografica": "Punto geoespacial por estación",
        "variables_clave": "nombre_estacion, lat, lon, tipo (troncal/cable/portal), corredor",
        "por_que_importante": "Feature de accesibilidad. Nodos de alta afluencia concentran ciertos tipos de crimen.",
        "valor_modelo": "Feature: distancia_transmilenio_mas_cercano por UPZ.",
        "licencia": "Pública",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "Infraestructura estable, pocas estaciones nuevas por año.",
        "volumen_estimado": "~150 estaciones troncales",
    },
    {
        "id": 10,
        "nombre": "Cámaras Salvavidas Bogotá (SDM)",
        "organizacion": "Secretaría Distrital de Movilidad (SDM)",
        "portal": "datos.movilidadbogota.gov.co",
        "plataforma": "ArcGIS Hub / ArcGIS REST",
        "tipo_acceso": "ArcGIS REST o descarga directa",
        "endpoint_url": "https://datos.movilidadbogota.gov.co/maps/camaras-salvavidas-bogota",
        "resource_id": "N/A (ArcGIS Hub)",
        "frecuencia_declarada": "Según necesidad",
        "ultima_actualizacion_real": "Verificar al descargar",
        "periodo_cobertura": "Actual",
        "granularidad_temporal": "Estática (infraestructura)",
        "granularidad_geografica": "Punto geoespacial por cámara",
        "variables_clave": "nombre_camara, lat, lon, tipo_control, corredor_vial, estado",
        "por_que_importante": "Feature de vigilancia tecnológica en corredores viales. Solo ubicación, NO el video.",
        "valor_modelo": "Feature: densidad_camaras_trafico por UPZ.",
        "licencia": "Datos abiertos SDM",
        "estado_url": None,
        "stale": "Verificar",
        "advertencias_calidad": "NO son cámaras CCTV de seguridad. Solo cámaras de tráfico. ArcGIS REST puede requerir token en algunos endpoints.",
        "volumen_estimado": "~200-500 cámaras",
    },
    {
        "id": 11,
        "nombre": "Siniestralidad BD — Accidentes Viales Bogotá (SDM diaria)",
        "organizacion": "Secretaría Distrital de Movilidad (SDM)",
        "portal": "datos.movilidadbogota.gov.co",
        "plataforma": "ArcGIS REST (actualización diaria desde SIGAT)",
        "tipo_acceso": "ArcGIS REST",
        "endpoint_url": "https://datos.movilidadbogota.gov.co/maps/ea243e7de8e846c8bd27e47c08771d66",
        "resource_id": "N/A (ArcGIS Hub)",
        "frecuencia_declarada": "DIARIA",
        "ultima_actualizacion_real": "Diaria (fuente SIGAT activa)",
        "periodo_cobertura": "Desde 2007",
        "granularidad_temporal": "Diaria",
        "granularidad_geografica": "Punto geoespacial por accidente",
        "variables_clave": "lat, lon, severidad, causa, tipo_vehiculo, actor_vial, fecha, hora, via",
        "por_que_importante": "Accidentalidad por UPZ como proxy de desorden vial. Crimen y accidentes comparten patrones espaciales.",
        "valor_modelo": "Feature: accidentalidad_upz — complementa crimen con proxy de desorden.",
        "licencia": "Datos abiertos SDM",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "ArcGIS REST puede paginar. Historial completo (2007-2026) puede ser varios GB.",
        "volumen_estimado": ">1M registros (desde 2007)",
    },
    {
        "id": 12,
        "nombre": "Estratificación Manzana Bogotá D.C. — SDP",
        "organizacion": "Secretaría Distrital de Planeación (SDP)",
        "portal": "datosabiertos.bogota.gov.co",
        "plataforma": "CKAN — descarga directa",
        "tipo_acceso": "Descarga GeoJSON / SHP ZIP",
        "endpoint_url": "https://datosabiertos.bogota.gov.co/dataset/55467552-0af4-4524-a390-a2956035744e/resource/29f2d770-bd5d-4450-9e95-8737167ba12f/download/manzanaestratificacion.json",
        "resource_id": "29f2d770-bd5d-4450-9e95-8737167ba12f",
        "frecuencia_declarada": "Según necesidad",
        "ultima_actualizacion_real": "20/11/2025",
        "periodo_cobertura": "Actual",
        "granularidad_temporal": "Estática",
        "granularidad_geografica": "Por manzana catastral",
        "variables_clave": "estrato (1-6), geometry, localidad",
        "por_que_importante": "OBLIGATORIA para análisis de sesgo algorítmico (Notebook 05). El jurado preguntará: '¿discrimina por estrato?'",
        "valor_modelo": "Feature: estrato_promedio_upz. Input para análisis de bias.",
        "licencia": "CC BY 4.0",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "Spatial join manzana→UPZ costoso — pre-agregar y guardar como Parquet.",
        "volumen_estimado": ">100K manzanas",
    },
    {
        "id": 13,
        "nombre": "Sensores de Aforo Vehicular y de Bicicletas (SDM)",
        "organizacion": "Secretaría Distrital de Movilidad (SDM)",
        "portal": "serviciosgis.catastrobogota.gov.co",
        "plataforma": "ArcGIS REST",
        "tipo_acceso": "ArcGIS REST",
        "endpoint_url": "https://serviciosgis.catastrobogota.gov.co/arcgis/rest/services/movilidad/controltransito/MapServer",
        "resource_id": "Layer 1 (aforo vehicular), Layer 2 (aforo bicicletas)",
        "frecuencia_declarada": "Infraestructura estable",
        "ultima_actualizacion_real": "Verificar",
        "periodo_cobertura": "Actual",
        "granularidad_temporal": "Estática (ubicación sensores)",
        "granularidad_geografica": "Punto geoespacial por sensor",
        "variables_clave": "lat, lon, tipo_sensor, corredor_vial",
        "por_que_importante": "Feature de densidad de tráfico por corredor → proxy de flujo peatonal y vehicular.",
        "valor_modelo": "Feature: densidad_sensores_aforo por UPZ.",
        "licencia": "Pública",
        "estado_url": None,
        "stale": "Verificar",
        "advertencias_calidad": "Solo puntos de sensores, no conteos en tiempo real.",
        "volumen_estimado": "~200-300 sensores",
    },
    {
        "id": 14,
        "nombre": "Censo Nacional 2018 + Encuesta Multipropósito Bogotá 2021 (DANE)",
        "organizacion": "DANE",
        "portal": "microdatos.dane.gov.co",
        "plataforma": "Descarga directa (registro gratuito)",
        "tipo_acceso": "Descarga CSV/SAV con registro",
        "endpoint_url": "https://microdatos.dane.gov.co/index.php/catalog/643",
        "resource_id": "catalog/643 (Censo 2018) + catalog/565 (EMB 2021)",
        "frecuencia_declarada": "Decenal (censo) / 2-3 años (EMB)",
        "ultima_actualizacion_real": "2018 (censo), 2021 (EMB)",
        "periodo_cobertura": "2018-2021",
        "granularidad_temporal": "Punto en el tiempo",
        "granularidad_geografica": "UPZ (EMB tiene cod_upz)",
        "variables_clave": "poblacion, hacinamiento, nivel_educativo, victimizacion_percibida (EMB), cod_upz",
        "por_que_importante": "Variables socioeconómicas de fondo: densidad, hacinamiento, educación → predictores de crimen estructural.",
        "valor_modelo": "Features: densidad_poblacional_upz, indice_hacinamiento_upz.",
        "licencia": "Registro básico gratuito requerido",
        "estado_url": None,
        "stale": "⚠️ 2018-2021",
        "advertencias_calidad": "Datos desactualizados. Para población actual usar proyecciones DANE 2024.",
        "volumen_estimado": "Varios GB (censo completo)",
    },
    {
        "id": 15,
        "nombre": "Grupo de Uso Económico Predominante por Manzana (UAECD)",
        "organizacion": "Unidad Administrativa Especial de Catastro Distrital (UAECD)",
        "portal": "datosabiertos.bogota.gov.co",
        "plataforma": "CKAN — descarga directa",
        "tipo_acceso": "Descarga GeoJSON / SHP / GPKG",
        "endpoint_url": "https://datosabiertos.bogota.gov.co/dataset/grupo_uso-por-manzana",
        "resource_id": "grupo_uso-por-manzana",
        "frecuencia_declarada": "Diaria (metadato actualizado abril 2026)",
        "ultima_actualizacion_real": "01/01/2026",
        "periodo_cobertura": "2012–2026 (versiones históricas disponibles)",
        "granularidad_temporal": "Anual (versión por año)",
        "granularidad_geografica": "Por manzana catastral",
        "variables_clave": "uso_predominante (Residencia/Comercio y oficinas/Depósitos/Industrial/Dotacional), geometry",
        "por_que_importante": "Feature de densidad comercial sólida. Comercio y oficinas → proxy de bares/vida nocturna → predictor de crimen violento (PMC 2012).",
        "valor_modelo": "Feature: pct_manzanas_comercio_upz — reemplaza densidad comercial estimada.",
        "licencia": "CC BY 4.0",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "Spatial join manzana→UPZ costoso — pre-agregar y guardar como Parquet. No distingue subtipos de comercio.",
        "volumen_estimado": ">100K manzanas",
    },
    {
        "id": 16,
        "nombre": "Lesiones no fatales de causa externa — Medicina Legal",
        "organizacion": "Instituto Nacional de Medicina Legal y Ciencias Forenses",
        "portal": "www.datos.gov.co",
        "plataforma": "Socrata",
        "tipo_acceso": "API Socrata (sodapy)",
        "endpoint_url": "https://www.datos.gov.co/resource/79dd-d24f.json",
        "resource_id": "79dd-d24f",
        "frecuencia_declarada": "Mensual",
        "ultima_actualizacion_real": "Marzo 2026",
        "periodo_cobertura": "Enero 2024 – Marzo 2026",
        "granularidad_temporal": "Mensual agregado",
        "granularidad_geografica": "Localidad (no UPZ)",
        "variables_clave": "a_o_del_hecho, mes_del_hecho, rango_de_hora, localidad_del_hecho, escenario_del_hecho, circunstancia_del_hecho, mecanismo_causal, presunto_agresor",
        "por_que_importante": "Proxy de subregistro: violencias que llegan a urgencias pero no a policía. ratio lesiones_ML/delitos_formales por localidad.",
        "valor_modelo": "Feature de subregistro por localidad → agregar a UPZ via tabla de correspondencia.",
        "licencia": "Datos abiertos públicos",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "Sin App Token: límite 1000 filas/request. Versión 'preliminar' — puede tener correcciones retroactivas. Solo desde 2024.",
        "volumen_estimado": ">50K registros",
    },
    {
        "id": 17,
        "nombre": "Lesiones fatales de causa externa — Medicina Legal",
        "organizacion": "Instituto Nacional de Medicina Legal y Ciencias Forenses",
        "portal": "www.datos.gov.co",
        "plataforma": "Socrata",
        "tipo_acceso": "API Socrata (sodapy)",
        "endpoint_url": "https://www.datos.gov.co/resource/2kpj-cktv.json",
        "resource_id": "2kpj-cktv",
        "frecuencia_declarada": "Mensual",
        "ultima_actualizacion_real": "Ene 2023 – Jul 2024",
        "periodo_cobertura": "Multianual",
        "granularidad_temporal": "Mensual",
        "granularidad_geografica": "Municipal / Localidad (Bogotá)",
        "variables_clave": "fecha_hecho, localidad_del_hecho, causa_de_muerte, clase_de_accidente",
        "por_que_importante": "Cross-valida homicidios entre Policía Nacional (FUENTE 3) y Medicina Legal. La diferencia = indicador de calidad del dato.",
        "valor_modelo": "Validación cruzada de variable objetivo (homicidios).",
        "licencia": "Datos abiertos públicos",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "Mismo rate limiting que FUENTE 16. Solo nivel municipal/localidad.",
        "volumen_estimado": ">20K registros",
    },
    {
        "id": 18,
        "nombre": "Incautación de Estupefacientes — Policía Nacional",
        "organizacion": "Policía Nacional de Colombia",
        "portal": "www.datos.gov.co",
        "plataforma": "Socrata",
        "tipo_acceso": "API Socrata (sodapy)",
        "endpoint_url": "https://www.datos.gov.co/resource/kk69-w2jj.json",
        "resource_id": "kk69-w2jj",
        "frecuencia_declarada": "Según reporte policial",
        "ultima_actualizacion_real": "Datos 2024 (verificar)",
        "periodo_cobertura": "Multianual",
        "granularidad_temporal": "Por evento",
        "granularidad_geografica": "Municipal",
        "variables_clave": "departamento, municipio, codigo_dane, clase_bien (tipo droga), fecha_hecho, cantidad",
        "por_que_importante": "Proxy de presencia de economías ilegales — microtráfico por localidad/municipio.",
        "valor_modelo": "Variable contextual: actividad antidroga Bogotá por año. No feature directa UPZ.",
        "licencia": "Datos abiertos públicos",
        "estado_url": None,
        "stale": "Verificar",
        "advertencias_calidad": "Solo nivel municipal (no UPZ). Refleja acción policial, no distribución real del mercado. Filtrar: municipio='BOGOTA D.C.'",
        "volumen_estimado": ">50K incautaciones",
    },
    {
        "id": 19,
        "nombre": "Points of Interest (POI) — OpenStreetMap / Overpass API",
        "organizacion": "OpenStreetMap Foundation",
        "portal": "openstreetmap.org / overpass-api.de",
        "plataforma": "REST API (Overpass QL) + overpy>=0.6.0",
        "tipo_acceso": "API REST gratuita (Overpass QL)",
        "endpoint_url": "https://overpass-api.de/api/interpreter",
        "resource_id": "N/A (consultas Overpass QL)",
        "frecuencia_declarada": "Actualización continua (crowdsourced)",
        "ultima_actualizacion_real": "Siempre actual",
        "periodo_cobertura": "Datos actuales",
        "granularidad_temporal": "Punto en el tiempo (estable)",
        "granularidad_geografica": "Punto geoespacial → spatial join con UPZ",
        "variables_clave": "amenity (bar/pub/nightclub/atm/bank/park/school), shop, lat, lon",
        "por_que_importante": "Feature de atractores de crimen. Densidad de bares, ATMs, parques por UPZ. Respaldado por Springer 2024 y PMC 2012.",
        "valor_modelo": "Features: n_bares_upz, n_atms_upz, densidad_vida_nocturna_upz, distancia_atm_mas_cercano.",
        "licencia": "ODbL — uso libre comercial y académico",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "Cobertura OSM en Bogotá puede ser incompleta en zonas periféricas. Verificar vs datos SDDE.",
        "volumen_estimado": "Variable (miles de POIs en Bogotá)",
    },
    {
        "id": 20,
        "nombre": "VIIRS Nighttime Lights — NASA/NOAA (Colorado School of Mines)",
        "organizacion": "NASA / NOAA / Earth Observation Group",
        "portal": "eogdata.mines.edu/products/vnl",
        "plataforma": "Descarga GeoTIFF (registro gratuito NASA Earthdata)",
        "tipo_acceso": "Descarga directa GeoTIFF (registro gratuito)",
        "endpoint_url": "https://eogdata.mines.edu/products/vnl/",
        "resource_id": "N/A (archivos mensuales GeoTIFF)",
        "frecuencia_declarada": "Mensual (composites cloud-free)",
        "ultima_actualizacion_real": "Composites mensuales actualizados regularmente",
        "periodo_cobertura": "Desde 2012 (Suomi NPP)",
        "granularidad_temporal": "Mensual",
        "granularidad_geografica": "~750m por píxel → agregar por UPZ (zonal stats)",
        "variables_clave": "radiancia_nocturna (nW/cm²/sr) — proxy de iluminación real desde satélite",
        "por_que_importante": "Reemplaza limitación de FUENTE 8 stale. Mide iluminación REAL vs declarada. Correlación negativa con crimen (ScienceDirect 2025).",
        "valor_modelo": "Feature: luz_nocturna_media_upz. Detecta luminarias apagadas vs declaradas.",
        "licencia": "NASA Open Data Policy — completamente libre",
        "estado_url": None,
        "stale": "No",
        "advertencias_calidad": "GeoTIFF globales (~100-300MB). Requiere rasterio + rasterstats.zonal_stats(). Composites ya filtran nubes.",
        "volumen_estimado": "~200MB por composite mensual",
    },
]

# ─────────────────────────────────────────────────────────────────────────────
# Fuentes descartadas
# ─────────────────────────────────────────────────────────────────────────────

DESCARTADAS = [
    {
        "fuente": "SIEDCO 2bxu-b96f",
        "razon": "404 en CKAN Bogotá — no existe como dataset tabular independiente",
        "alternativa": "FUENTE 2 (Delito de Alto Impacto — mismo origen, más reciente y con coordenadas)",
    },
    {
        "fuente": "Cámaras CCTV de seguridad (C4)",
        "razon": "Ley 1581/2012 — ubicaciones e imágenes no son datos abiertos",
        "alternativa": "FUENTE 10 (cámaras de tráfico SDM — ubicaciones SÍ son abiertas)",
    },
    {
        "fuente": "Siniestros Viales — CKAN Bogotá",
        "razon": "STALE — última actualización octubre 2021 (anual, desactualizada)",
        "alternativa": "FUENTE 11 (SDM versión diaria via ArcGIS REST SIGAT)",
    },
    {
        "fuente": "Tasa Desempleo SDDE (portal CKAN)",
        "razon": "Sin actualizar desde diciembre 2023 — desactualizado para el modelo",
        "alternativa": "FUENTE 14 (DANE EMB 2021 con cod_upz)",
    },
    {
        "fuente": "Banco de la República (series de criminalidad)",
        "razon": "Solo nivel departamental, no UPZ. No aporta features del modelo.",
        "alternativa": "Usar para contexto en presentación oral, no como feature",
    },
    {
        "fuente": "INPEC (datos penitenciarios)",
        "razon": "Desagregación nacional — no UPZ, no Bogotá",
        "alternativa": "No aplica para el modelo UPZ",
    },
    {
        "fuente": "SIMUR (movilidad urbana)",
        "razon": "Datos fragmentados, acceso inconsistente",
        "alternativa": "FUENTE 13 (sensores SDM via ArcGIS REST directo)",
    },
    {
        "fuente": "Redes sociales — scraping (Twitter/X, Instagram)",
        "razon": "Viola reglas del concurso si scraping. X API ahora $100+/mes. Ley 1581/2012. Solo 1-3% tweets geolocalizados en Colombia.",
        "alternativa": "Google Trends como proxy de percepción (API oficial jul 2025 — opcional, 2da prioridad)",
    },
    {
        "fuente": "APIs de pago (Google Maps Platform)",
        "razon": "Viola restricción del concurso de APIs de pago como fuente principal",
        "alternativa": "Open-Meteo, IDECA, OSM Overpass — todas gratuitas",
    },
    {
        "fuente": "IDU Obras en construcción activas",
        "razon": "No existe API descargable — solo mapa interactivo en bogota.gov.co/mapa-obras-idu sin endpoint público",
        "alternativa": "IDU Estado Superficial vial (stale 2023) como proxy de deterioro urbano si es necesario",
    },
    {
        "fuente": "Violencia Intrafamiliar — SDS",
        "razon": "Solo tasa por 100K hab para Bogotá completo, anual. Sin desagregación UPZ ni localidad.",
        "alternativa": "FUENTE 16 (Medicina Legal — lesiones no fatales, con localidad_del_hecho, 40 columnas)",
    },
    {
        "fuente": "Google Street View (visual disorder scoring)",
        "razon": "Requiere Google Maps API de pago para 150K+ imágenes para cubrir Bogotá. Viola restricción del concurso.",
        "alternativa": "FUENTE 19 (OSM POI) como proxy estructural + FUENTE 20 (VIIRS) para iluminación real",
    },
    {
        "fuente": "Banco de Imágenes CCTV / Video analítico",
        "razon": "Video de cámaras NO es dato abierto (Ley 1581/2012). Solo las ubicaciones son públicas.",
        "alternativa": "Ubicaciones vía FUENTE 10 (cámaras SDM). Video analítico queda como roadmap futuro.",
    },
]

# ─────────────────────────────────────────────────────────────────────────────
# IDs de fuentes geoespaciales (con polígonos/puntos en UPZ)
# ─────────────────────────────────────────────────────────────────────────────

GEO_IDS = {6, 7, 9, 10, 11, 12, 13, 15}

GEO_DETAIL = {
    6: {
        "crs": "EPSG:4326 (WGS84)",
        "features": "112 polígonos (verificar — no 117 UPL)",
        "formato": "GeoJSON / SHP ZIP",
        "spatial_join": "upz = gpd.read_file(url); crimen_gdf = gpd.GeoDataFrame(df, geometry=...); gpd.sjoin(upz, crimen_gdf, how='left', predicate='contains')",
        "por_que_importante": "Es la CAPA BASE de todo el modelo. Sin estos polígonos no se pueden hacer spatial joins, no existe el mapa del dashboard y no hay unidad de análisis. Cada una de las 112 UPZs es una fila del dataset de entrenamiento.",
        "feature_modelo": "No genera features directamente — ES la unidad de análisis. Todos los demás datasets se agregan a nivel UPZ usando esta geometría como referencia.",
    },
    7: {
        "crs": "EPSG:4326",
        "features": "~1200 cuadrantes",
        "formato": "GeoJSON / SHP ZIP",
        "spatial_join": "gpd.sjoin(upz, cuadrantes) → groupby('codigo_upz').size() / upz['area_km2'] → feature: cuadrantes_por_km2_upz",
        "por_que_importante": "Mide la cobertura policial real de cada UPZ. Las zonas con menos cuadrantes por km² tienen menor presencia policial y tienden a mayor criminalidad. También es input de la capa prescriptiva: cuando el riesgo es 'temporal', la recomendación es reforzar cuadrantes (intervención MEBOG/SIJIN).",
        "feature_modelo": "cuadrantes_por_km2_upz — densidad de cobertura policial. Feature urbanística del modelo predictivo y variable de decisión en la capa prescriptiva.",
    },
    9: {
        "crs": "EPSG:4326",
        "features": "~150 estaciones troncales",
        "formato": "CSV con lat/lon / GeoJSON",
        "spatial_join": "estaciones_gdf = gpd.GeoDataFrame(df, geometry=gpd.points_from_xy(df.lon, df.lat)); gpd.sjoin(upz, estaciones_gdf) → count() por UPZ",
        "por_que_importante": "Los nodos de alta afluencia de TransMilenio (portales, estaciones de integración) concentran hurto a personas y riñas. La distancia de una UPZ a la estación más cercana captura su nivel de exposición a flujos masivos de personas.",
        "feature_modelo": "n_estaciones_tm_upz (conteo de estaciones dentro de la UPZ) + distancia_tm_mas_cercana (para UPZs sin estación propia). Feature de movilidad y exposición.",
    },
    10: {
        "crs": "EPSG:4326",
        "features": "~200-500 cámaras",
        "formato": "GeoJSON (ArcGIS REST)",
        "spatial_join": "gpd.sjoin(upz, camaras_gdf) → count() por UPZ → feature: n_camaras_trafico_upz",
        "por_que_importante": "La presencia de cámaras de control de velocidad en corredores viales funciona como proxy de vigilancia tecnológica en la zona. Las UPZs con más cámaras visibles tienen un efecto disuasorio sobre ciertos tipos de crimen de oportunidad.",
        "feature_modelo": "n_camaras_trafico_upz — densidad de vigilancia tecnológica vial. Feature urbanística del modelo.",
    },
    11: {
        "crs": "EPSG:4326",
        "features": ">1M registros desde 2007",
        "formato": "GeoJSON (ArcGIS REST — paginado)",
        "spatial_join": "gpd.sjoin(upz, siniestros_gdf) → groupby(['codigo_upz', 'anio']).size() → feature: accidentalidad_anual_upz",
        "por_que_importante": "Los accidentes de tráfico y el crimen urbano comparten causas estructurales: corredores mal señalizados, ausencia de control, alta velocidad nocturna, zonas con bajo orden. La accidentalidad por UPZ es un proxy de 'desorden urbano' que el modelo puede usar para capturar riesgo latente no visible en datos de crimen puro.",
        "feature_modelo": "accidentalidad_anual_upz — proxy de desorden vial y urbano. Feature complementaria que añade señal sobre el estado de orden público de la zona más allá del crimen reportado.",
    },
    12: {
        "crs": "EPSG:4326",
        "features": ">100K manzanas",
        "formato": "GeoJSON / SHP ZIP / GPKG",
        "spatial_join": "gpd.sjoin(manzanas, upz) → groupby('codigo_upz')['estrato'].mean() → feature: estrato_promedio_upz",
        "por_que_importante": "El estrato socioeconómico es una de las variables más correlacionadas con el crimen en ciudades colombianas. Pero más importante aún: es OBLIGATORIA para el análisis de sesgo algorítmico del Notebook 05. El jurado preguntará '¿el modelo discrimina por estrato?' — esta capa permite responder con datos.",
        "feature_modelo": "estrato_promedio_upz — contexto socioeconómico de la zona. Doble uso: feature predictiva del modelo Y variable de análisis de equidad/bias en Notebook 05.",
    },
    13: {
        "crs": "EPSG:4326",
        "features": "~200-300 sensores",
        "formato": "GeoJSON (ArcGIS REST — Layer 1 vehículos, Layer 2 bicicletas)",
        "spatial_join": "gpd.sjoin(upz, sensores_gdf) → count() por UPZ → feature: n_sensores_aforo_upz",
        "por_que_importante": "El volumen de tráfico vehicular y peatonal de una zona determina su exposición al crimen de oportunidad. Los sensores de aforo están ubicados en los corredores de mayor flujo — su densidad en una UPZ es un proxy de la intensidad de circulación que no está capturada por otras fuentes.",
        "feature_modelo": "n_sensores_aforo_upz — proxy de intensidad de tráfico en corredores principales. Feature de movilidad que complementa la distancia a TransMilenio.",
    },
    15: {
        "crs": "EPSG:4326 (escala 1:1000)",
        "features": ">100K manzanas",
        "formato": "GeoJSON / SHP / GPKG",
        "spatial_join": "gpd.sjoin(manzanas, upz) → groupby('codigo_upz')['uso_predominante'].value_counts(normalize=True) → pivot → pct_comercio_upz",
        "por_que_importante": "La densidad de manzanas comerciales en una UPZ es el mejor proxy disponible de vida nocturna, bares y establecimientos de consumo — que múltiples estudios internacionales (PMC 2012, Springer 2024) identifican como los predictores más fuertes de crimen violento. Reemplaza la estimación indirecta de 'densidad comercial' que antes era la variable menos justificada del modelo.",
        "feature_modelo": "pct_manzanas_comercio_upz — porcentaje de manzanas con uso 'Comercio y oficinas'. Feature urbanística con respaldo bibliográfico sólido para predecir crimen violento nocturno.",
    },
}

# ─────────────────────────────────────────────────────────────────────────────
# Impacto si la fuente está stale
# ─────────────────────────────────────────────────────────────────────────────

IMPACTO_STALE = {
    1: "Bajo — datos agregados mensuales, el patrón no cambia significativamente",
    2: "Bajo — infraestructura física de crimen cambia lentamente entre semestres",
    3: "Bajo — contexto histórico, no variable principal del modelo UPZ",
    4: "Bajo — contexto histórico, no variable principal del modelo UPZ",
    5: "Nulo — siempre actual (API en tiempo real)",
    6: "Nulo — geometría de UPZs no cambia salvo reforma administrativa",
    7: "Bajo — cuadrantes se reorganizan ocasionalmente (anual)",
    8: "Medio — documentar en Notebook 05. Compensado por FUENTE 20 (VIIRS iluminación real).",
    9: "Bajo — red TransMilenio crece lentamente (1-2 estaciones/año)",
    10: "Bajo — infraestructura de cámaras de tráfico es estable",
    11: "Nulo — fuente SIGAT diaria activa",
    12: "Bajo — estratificación cambia por decreto (infrecuente, ~cada 5 años)",
    13: "Bajo — sensores de aforo se reemplazan raramente",
    14: "Medio — población cambia, pero proyecciones DANE 2024 cubren la brecha",
    15: "Bajo — uso del suelo cambia lentamente, versión 2026 disponible",
    16: "Bajo — solo desde 2024, periodo corto pero activo y actualizado",
    17: "Bajo — para validación cruzada, no feature principal",
    18: "Bajo — contextual, no feature directa a nivel UPZ",
    19: "Nulo — siempre actual (crowdsourced OSM, actualización continua)",
    20: "Nulo — composites mensuales actualizados regularmente por EOG/NASA",
}

# ─────────────────────────────────────────────────────────────────────────────
# Validación de URLs
# ─────────────────────────────────────────────────────────────────────────────


def validar_todas() -> None:
    print("Validando URLs de 20 fuentes...\n")
    ok = 0
    for f in FUENTES:
        url = f["endpoint_url"]
        status, t = check_url(url)
        if 200 <= status < 400:
            f["estado_url"] = f"OK {status}"
            simbolo = "✅"
            ok += 1
        else:
            f["estado_url"] = f"ERROR {status}"
            simbolo = "❌"
        print(f"[{f['id']:2d}] {simbolo} {f['nombre'][:48]:<48}  HTTP {status:4d}  {t:.1f}s")
    print(f"\n{ok}/20 fuentes responden correctamente.")


# ─────────────────────────────────────────────────────────────────────────────
# Estilos Excel
# ─────────────────────────────────────────────────────────────────────────────

HEADER_FILL = PatternFill("solid", fgColor="1F497D")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL = PatternFill("solid", fgColor="E8F0FE")
STALE_FILL = PatternFill("solid", fgColor="FFE0B2")
GREEN_FILL = PatternFill("solid", fgColor="E8F5E9")
YELLOW_FILL = PatternFill("solid", fgColor="FFF9C4")
RED_FILL = PatternFill("solid", fgColor="FFCDD2")
WRAP_TOP = Alignment(wrap_text=True, vertical="top")


def _autofit(ws) -> None:
    for col_cells in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 55)


def _write_header(ws, headers: list[str]) -> None:
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="center")
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 28


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 1: Catálogo
# ─────────────────────────────────────────────────────────────────────────────

COLS_CATALOGO = [
    "#", "Nombre", "Organización", "Portal", "Plataforma", "Tipo_acceso",
    "Endpoint_URL_verificada", "Resource_ID", "Frecuencia_declarada",
    "Ultima_actualizacion_real", "Período_cobertura", "Granularidad_temporal",
    "Granularidad_geografica", "Variables_clave", "Por_que_es_importante",
    "Valor_para_modelo", "Licencia", "Estado_URL", "Stale",
    "Advertencias_calidad", "Volumen_estimado",
]

KEYS_CATALOGO = [
    "id", "nombre", "organizacion", "portal", "plataforma", "tipo_acceso",
    "endpoint_url", "resource_id", "frecuencia_declarada",
    "ultima_actualizacion_real", "periodo_cobertura", "granularidad_temporal",
    "granularidad_geografica", "variables_clave", "por_que_importante",
    "valor_modelo", "licencia", "estado_url", "stale",
    "advertencias_calidad", "volumen_estimado",
]


def _fill_catalogo(ws) -> None:
    _write_header(ws, COLS_CATALOGO)
    for row_i, f in enumerate(FUENTES, 2):
        stale_str = str(f.get("stale", ""))
        row_fill = STALE_FILL if "SÍ" in stale_str else (ALT_FILL if row_i % 2 == 0 else None)
        for col_i, key in enumerate(KEYS_CATALOGO, 1):
            cell = ws.cell(row=row_i, column=col_i, value=f.get(key))
            cell.alignment = WRAP_TOP
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row_i].height = 55
    _autofit(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 2: Geoespaciales
# ─────────────────────────────────────────────────────────────────────────────

# Hoja 2 usa todas las columnas del Catálogo + 3 columnas espaciales al final
COLS_GEO = COLS_CATALOGO + [
    "CRS_esperado",
    "Num_features_esperadas",
    "Cómo_hacer_spatial_join_con_UPZ",
]

KEYS_GEO_BASE = KEYS_CATALOGO  # mismas claves que Catálogo


def _fill_geoespaciales(ws) -> None:
    _write_header(ws, COLS_GEO)
    row_i = 2
    for f in FUENTES:
        if f["id"] not in GEO_IDS:
            continue
        geo = GEO_DETAIL.get(f["id"], {})
        stale_str = str(f.get("stale", ""))
        row_fill = STALE_FILL if "SÍ" in stale_str else (ALT_FILL if row_i % 2 == 0 else None)

        # Para Por_que_es_importante y Valor_para_modelo usamos la versión
        # detallada de GEO_DETAIL si existe, sino la genérica de FUENTES
        f_geo = dict(f)
        if geo.get("por_que_importante"):
            f_geo["por_que_importante"] = geo["por_que_importante"]
        if geo.get("feature_modelo"):
            f_geo["valor_modelo"] = geo["feature_modelo"]

        # Columnas del Catálogo
        for col_i, key in enumerate(KEYS_GEO_BASE, 1):
            cell = ws.cell(row=row_i, column=col_i, value=f_geo.get(key))
            cell.alignment = WRAP_TOP
            if row_fill:
                cell.fill = row_fill

        # Columnas espaciales extra (al final)
        n = len(KEYS_GEO_BASE)
        for extra_col, val in enumerate([
            geo.get("crs", "EPSG:4326"),
            geo.get("features", ""),
            geo.get("spatial_join", ""),
        ], n + 1):
            cell = ws.cell(row=row_i, column=extra_col, value=val)
            cell.alignment = WRAP_TOP
            if row_fill:
                cell.fill = row_fill

        ws.row_dimensions[row_i].height = 100
        row_i += 1
    _autofit(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 3: Descartadas
# ─────────────────────────────────────────────────────────────────────────────

COLS_DESCARTADAS = ["Fuente_descartada", "Razón_de_descarte", "Alternativa"]


def _fill_descartadas(ws) -> None:
    _write_header(ws, COLS_DESCARTADAS)
    for row_i, d in enumerate(DESCARTADAS, 2):
        row_fill = ALT_FILL if row_i % 2 == 0 else None
        for col_i, key in enumerate(["fuente", "razon", "alternativa"], 1):
            cell = ws.cell(row=row_i, column=col_i, value=d[key])
            cell.alignment = WRAP_TOP
            if row_fill:
                cell.fill = row_fill
        ws.row_dimensions[row_i].height = 50
    _autofit(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Hoja 4: Frecuencias
# ─────────────────────────────────────────────────────────────────────────────

COLS_FREQ = [
    "#", "Nombre", "Frecuencia_declarada", "Ultima_actualizacion_real",
    "Stale", "Semaforo", "Impacto_si_stale",
]


def _semaforo(stale: str) -> tuple[str, PatternFill]:
    s = stale.upper()
    if "SÍ" in s or "STALE" in s:
        return "ROJO — STALE", RED_FILL
    if "VERIFICAR" in s or "2018" in s or "2021" in s:
        return "AMARILLO — Verificar", YELLOW_FILL
    return "VERDE — OK", GREEN_FILL


def _fill_frecuencias(ws) -> None:
    _write_header(ws, COLS_FREQ)
    for row_i, f in enumerate(FUENTES, 2):
        stale_str = str(f.get("stale", ""))
        semaforo_txt, sem_fill = _semaforo(stale_str)
        impacto = IMPACTO_STALE.get(f["id"], "")
        vals = [
            f["id"], f["nombre"], f["frecuencia_declarada"],
            f["ultima_actualizacion_real"], stale_str, semaforo_txt, impacto,
        ]
        for col_i, v in enumerate(vals, 1):
            cell = ws.cell(row=row_i, column=col_i, value=v)
            cell.alignment = WRAP_TOP
            cell.fill = sem_fill
        ws.row_dimensions[row_i].height = 45
    _autofit(ws)


# ─────────────────────────────────────────────────────────────────────────────
# Generación del Excel
# ─────────────────────────────────────────────────────────────────────────────


def generar_excel(output_path: str = "fuentes_validadas.xlsx") -> None:
    wb = openpyxl.Workbook()

    ws1 = wb.active
    ws1.title = "Catálogo"
    _fill_catalogo(ws1)

    ws2 = wb.create_sheet("Geoespaciales")
    _fill_geoespaciales(ws2)

    ws3 = wb.create_sheet("Descartadas")
    _fill_descartadas(ws3)

    ws4 = wb.create_sheet("Frecuencias")
    _fill_frecuencias(ws4)

    wb.save(output_path)
    print(f"\nExcel guardado: {output_path}")
    print(f"Hojas: {[ws.title for ws in wb.worksheets]}")


# ─────────────────────────────────────────────────────────────────────────────
# Entry point
# ─────────────────────────────────────────────────────────────────────────────

if __name__ == "__main__":
    validar_todas()
    generar_excel()
